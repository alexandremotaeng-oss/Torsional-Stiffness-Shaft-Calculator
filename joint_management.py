import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import os

NUM_SECTIONS = 10   # number of diameter/length section pairs per joint

COLUMNS = (
    "joint_name",
    "offset",
    *[c for i in range(1, NUM_SECTIONS + 1) for c in (f"section_{i}", f"length_{i}")],
)

# Fallback English headers used only when no translator is supplied
COLUMN_HEADERS_FALLBACK = {
    "joint_name": "Joint Name",
    "offset": "Offset (mm)",
    **{
        c: lbl
        for i in range(1, NUM_SECTIONS + 1)
        for c, lbl in [
            (f"section_{i}", f"Section {i} Ø (mm)"),
            (f"length_{i}",  f"Length {i} (mm)"),
        ]
    },
}

COLUMN_WIDTHS = {
    "joint_name": 140,
    "offset": 90,
    **{k: 85 for k in COLUMN_HEADERS_FALLBACK if k not in ("joint_name", "offset")},
}


class JointManagementWindow(tk.Toplevel):
    def __init__(self, parent, data_dir="data", translator=None):
        super().__init__(parent)
        self.data_dir = data_dir
        self.translator = translator
        self.title(self._t("joint_management_window_title"))
        self.resizable(True, True)

        self._build_ui()
        self.load_joints()
        self.update_ui()

        # Auto-fit to content; keep user able to resize freely.
        self.update_idletasks()
        w = 900
        h = max(self.winfo_reqheight(), 400)
        self.geometry(f"{w}x{h}")
        self.minsize(500, 250)

    def _t(self, key):
        if self.translator:
            return self.translator.translate(key)
        return COLUMN_HEADERS_FALLBACK.get(key, key)

    def _default_new_row(self):
        return (self._t("new_joint_default_name"),) + ("0.00",) * (1 + NUM_SECTIONS * 2)

    def update_ui(self):
        """Refresh all translatable text — called on creation and after language change."""
        self.title(self._t("joint_management_window_title"))
        self.new_button.config(text=self._t("new"))
        self.edit_button.config(text=self._t("edit"))
        self.save_button.config(text=self._t("save"))
        self.delete_button.config(text=self._t("delete"))
        self.export_button.config(text=self._t("export_list"))
        self.import_button.config(text=self._t("import_list"))
        # Refresh treeview column headers
        for col in COLUMNS:
            self.tree.heading(col, text=self._t(col))

    def _build_ui(self):
        # ── TOOLBAR ────────────────────────────────────────────────────
        toolbar = ttk.Frame(self)
        toolbar.pack(side="top", fill="x", padx=8, pady=6)

        btn_frame = ttk.Frame(toolbar)
        btn_frame.pack(anchor="center")

        self.new_button = ttk.Button(
            btn_frame, text=self._t("new"), command=self.new_joint, width=12
        )
        self.edit_button = ttk.Button(
            btn_frame, text=self._t("edit"), command=self.edit_joints, width=12
        )
        self.save_button = ttk.Button(
            btn_frame, text=self._t("save"), command=self.save_joints, width=12
        )
        self.delete_button = ttk.Button(
            btn_frame, text=self._t("delete"), command=self.delete_joint, width=12
        )
        self.export_button = ttk.Button(
            btn_frame, text=self._t("export_list"), command=self.export_list, width=14
        )
        self.import_button = ttk.Button(
            btn_frame, text=self._t("import_list"), command=self.import_list, width=14
        )

        for col, btn in enumerate([
            self.new_button, self.edit_button, self.save_button, self.delete_button,
            self.export_button, self.import_button,
        ]):
            btn.grid(row=0, column=col, padx=8, pady=2)

        # ── TREEVIEW + SCROLLBARS ──────────────────────────────────────
        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        self.tree = ttk.Treeview(tree_frame, columns=COLUMNS, show="headings",
                                  selectmode="browse")

        for col in COLUMNS:
            self.tree.heading(col, text=self._t(col), anchor="center")
            self.tree.column(col, width=COLUMN_WIDTHS[col], anchor="center",
                             minwidth=60, stretch=True)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        self.tree.bind("<Double-1>", self.on_double_click)

    # ------------------------------------------------------------------
    # Inline editing
    # ------------------------------------------------------------------
    def on_double_click(self, event):
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        item = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if not item or not col:
            return

        col_index = int(col.replace("#", "")) - 1
        bbox = self.tree.bbox(item, col)
        if not bbox:
            return
        x, y, width, height = bbox

        values = self.tree.item(item, "values")
        current_val = values[col_index] if col_index < len(values) else ""

        entry_var = tk.StringVar(value=current_val)
        entry = ttk.Entry(self.tree, textvariable=entry_var, justify="center")
        entry.place(x=x, y=y, width=width, height=height)
        entry.focus_set()
        entry.select_range(0, tk.END)

        def commit(e=None):
            new_val = entry_var.get()
            vals = list(self.tree.item(item, "values"))
            # Validate numeric columns (all except first)
            if col_index > 0:
                try:
                    new_val = f"{float(new_val):.2f}"
                except ValueError:
                    new_val = "0.00"
            vals[col_index] = new_val
            self.tree.item(item, values=vals)
            entry.destroy()

        entry.bind("<Return>", commit)
        entry.bind("<Tab>", commit)
        entry.bind("<FocusOut>", commit)

    # ------------------------------------------------------------------
    # CRUD
    # ------------------------------------------------------------------
    def new_joint(self):
        iid = self.tree.insert("", "end", values=self._default_new_row())
        self.tree.selection_set(iid)
        self.tree.see(iid)

    def edit_joints(self):
        messagebox.showinfo(
            self._t("edit"),
            self._t("edit_hint")
        )

    def save_joints(self):
        joints = []
        for child in self.tree.get_children():
            values = self.tree.item(child, "values")
            try:
                offset = float(values[1]) if values[1] else 0.0
                sections = []
                for i in range(NUM_SECTIONS):
                    col_base = 2 + i * 2   # index into values tuple
                    sections.append({
                        "diameter": float(values[col_base]) if values[col_base] else 0.0,
                        "length":   float(values[col_base + 1]) if values[col_base + 1] else 0.0,
                    })
                joints.append({"name": values[0], "offset": offset, "sections": sections})
            except (IndexError, ValueError) as exc:
                messagebox.showerror(self._t("error"), f"{self._t('invalid_data')}: {exc}")
                return

        path = os.path.join(self.data_dir, "joints.json")
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(joints, f, indent=4, ensure_ascii=False)
            messagebox.showinfo(self._t("save"), self._t("saved_joints_ok"))
        except Exception as exc:
            messagebox.showerror(self._t("error"), f"{self._t('save_error')}: {exc}")

    def delete_joint(self):
        selected = self.tree.selection()
        if not selected:
            return
        if messagebox.askyesno(self._t("delete"), self._t("delete_confirm_joint")):
            self.tree.delete(selected)

    # ------------------------------------------------------------------
    # Export / Import helpers
    # ------------------------------------------------------------------
    def _tree_to_rows(self):
        """Return (header_row, [data_rows]) from the treeview — plain lists of str."""
        headers = [self._t(col) for col in COLUMNS]
        rows = []
        for child in self.tree.get_children():
            rows.append(list(self.tree.item(child, "values")))
        return headers, rows

    def _rows_to_joints(self, headers, rows):
        """
        Parse a list of string rows (exported format) back into joint dicts.
        Returns list of dicts or raises ValueError with a descriptive message.
        """
        # Validate header count
        if len(headers) != len(COLUMNS):
            raise ValueError(
                self._t("import_bad_format").format(
                    expected=len(COLUMNS), got=len(headers)
                )
            )

        joints = []
        for row_idx, row in enumerate(rows, start=2):   # row 1 = header
            if len(row) < len(COLUMNS):
                row = row + ["0.00"] * (len(COLUMNS) - len(row))

            name   = str(row[0]).strip()
            if not name:
                continue   # skip blank rows

            try:
                offset = float(row[1])
            except (ValueError, TypeError):
                raise ValueError(
                    self._t("import_bad_cell").format(row=row_idx, col=self._t("offset"))
                )

            sections = []
            for i in range(NUM_SECTIONS):
                base = 2 + i * 2
                try:
                    d = float(row[base])
                    l = float(row[base + 1])
                except (ValueError, TypeError):
                    raise ValueError(
                        self._t("import_bad_cell").format(
                            row=row_idx,
                            col=f"section_{i+1}/length_{i+1}"
                        )
                    )
                sections.append({"diameter": d, "length": l})
            joints.append({"name": name, "offset": offset, "sections": sections})
        return joints

    # ── EXPORT ────────────────────────────────────────────────────────
    def export_list(self):
        if not self.tree.get_children():
            messagebox.showinfo(self._t("export_list"), self._t("export_empty"))
            return

        path = filedialog.asksaveasfilename(
            title=self._t("export_list"),
            defaultextension=".xlsx",
            filetypes=[
                ("Excel Workbook", "*.xlsx"),
                ("Tab-separated Text", "*.txt"),
            ],
            initialfile="joints_export",
        )
        if not path:
            return

        headers, rows = self._tree_to_rows()
        ext = os.path.splitext(path)[1].lower()

        try:
            if ext == ".xlsx":
                self._export_xlsx(path, headers, rows)
            else:
                self._export_txt(path, headers, rows)
            messagebox.showinfo(
                self._t("export_list"),
                self._t("export_ok").format(path=path),
            )
        except Exception as exc:
            messagebox.showerror(self._t("error"), f"{self._t('save_error')}: {exc}")

    def _export_xlsx(self, path, headers, rows):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Joints"

        # Header row style
        hdr_font  = Font(bold=True, color="FFFFFF", size=10)
        hdr_fill  = PatternFill("solid", fgColor="0D2540")
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin      = Side(style="thin", color="C8D8EA")
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)

        for c_idx, hdr in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c_idx, value=hdr)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align
            cell.border    = border

        # Data rows
        row_fill_even = PatternFill("solid", fgColor="F0F5FB")
        data_align    = Alignment(horizontal="center", vertical="center")

        for r_idx, row in enumerate(rows, start=2):
            fill = row_fill_even if r_idx % 2 == 0 else PatternFill()
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.alignment = data_align
                cell.border    = border
                if fill.fill_type:
                    cell.fill = fill

        # Column widths
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 12
        for col_letter in [ws.cell(row=1, column=c).column_letter
                           for c in range(3, len(headers) + 1)]:
            ws.column_dimensions[col_letter].width = 13
        ws.row_dimensions[1].height = 30

        wb.save(path)

    def _export_txt(self, path, headers, rows):
        """Tab-separated, compatible with Excel paste."""
        with open(path, "w", encoding="utf-8-sig", newline="\r\n") as f:
            f.write("\t".join(headers) + "\r\n")
            for row in rows:
                f.write("\t".join(str(v) for v in row) + "\r\n")

    # ── IMPORT ────────────────────────────────────────────────────────
    def import_list(self):
        path = filedialog.askopenfilename(
            title=self._t("import_list"),
            filetypes=[
                ("Excel / Text", "*.xlsx *.txt"),
                ("Excel Workbook", "*.xlsx"),
                ("Tab-separated Text", "*.txt"),
            ],
        )
        if not path:
            return

        ext = os.path.splitext(path)[1].lower()
        try:
            if ext == ".xlsx":
                headers, rows = self._read_xlsx(path)
            else:
                headers, rows = self._read_txt(path)
        except Exception as exc:
            messagebox.showerror(self._t("error"), f"{self._t('import_bad_format_title')}: {exc}")
            return

        # Parse rows → joint dicts (validates format)
        try:
            imported_joints = self._rows_to_joints(headers, rows)
        except ValueError as exc:
            messagebox.showerror(self._t("import_bad_format_title"), str(exc))
            return

        if not imported_joints:
            messagebox.showinfo(self._t("import_list"), self._t("import_empty"))
            return

        # Build current DB dict keyed by name
        current_db: dict[str, dict] = {}
        for child in self.tree.get_children():
            vals = self.tree.item(child, "values")
            name = vals[0]
            current_db[name] = {"iid": child, "values": vals}

        added   = 0
        updated = 0
        skipped = 0

        for joint in imported_joints:
            name = joint["name"]
            # Build the treeview values tuple for this joint
            new_vals = [name, f"{joint['offset']:.2f}"]
            for sec in joint["sections"]:
                new_vals.append(f"{sec['diameter']:.2f}")
                new_vals.append(f"{sec['length']:.2f}")
            new_vals_tuple = tuple(new_vals)

            if name not in current_db:
                # New joint → add immediately
                self.tree.insert("", "end", values=new_vals_tuple)
                added += 1
            else:
                # Existing joint → compare
                existing_vals = current_db[name]["values"]
                if tuple(existing_vals) != new_vals_tuple:
                    answer = messagebox.askyesno(
                        self._t("import_diff_title"),
                        self._t("import_diff_msg").format(name=name),
                    )
                    if answer:
                        iid = current_db[name]["iid"]
                        self.tree.item(iid, values=new_vals_tuple)
                        updated += 1
                    else:
                        skipped += 1

        messagebox.showinfo(
            self._t("import_list"),
            self._t("import_done").format(added=added, updated=updated, skipped=skipped),
        )

    def _read_xlsx(self, path):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not all_rows:
            raise ValueError(self._t("import_bad_format_title"))
        headers = [str(c).strip() if c is not None else "" for c in all_rows[0]]
        rows    = [
            [str(c).strip() if c is not None else "0" for c in row]
            for row in all_rows[1:]
            if any(c not in (None, "", "None") for c in row)
        ]
        return headers, rows

    def _read_txt(self, path):
        """Read a tab-separated .txt file (UTF-8 or UTF-8-BOM)."""
        with open(path, "r", encoding="utf-8-sig") as f:
            lines = [ln.rstrip("\r\n") for ln in f if ln.strip()]
        if not lines:
            raise ValueError(self._t("import_bad_format_title"))
        headers = lines[0].split("\t")
        rows    = [ln.split("\t") for ln in lines[1:]]
        return headers, rows

    # ------------------------------------------------------------------
    def load_joints(self):
        path = os.path.join(self.data_dir, "joints.json")
        try:
            with open(path, "r", encoding="utf-8") as f:
                joints = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            joints = []
        except Exception as exc:
            messagebox.showerror(self._t("error"), f"{self._t('load_error')}: {exc}")
            return

        for joint in joints:
            values = [joint.get("name", ""), f"{float(joint.get('offset', 0.0)):.2f}"]
            for sec in joint.get("sections", []):
                values.append(f"{float(sec.get('diameter', 0)):.2f}")
                values.append(f"{float(sec.get('length', 0)):.2f}")
            # Pad to full column count if needed
            while len(values) < len(COLUMNS):
                values.append("0.00")
            self.tree.insert("", "end", values=tuple(values[:len(COLUMNS)]))
